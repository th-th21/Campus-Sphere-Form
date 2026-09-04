import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from io import BytesIO


# ============================================================
# CAMPUS SPHERE
# ============================================================

st.set_page_config(
    page_title="Campus Sphere",
    page_icon="💜",
    layout="wide"
)


# ============================================================
# SETTINGS
# ============================================================

DATA_FILE = Path("Campus_Sphere_Responses.xlsx")

# CHANGE THIS TO YOUR PRIVATE ADMIN PASSWORD
ADMIN_PASSWORD = "MyPrivatePassword2026"

PURPLE = "#4B248F"


# ============================================================
# OPTIONS
# ============================================================

ACADEMIC_OPTIONS = [
    "All",
    "Class Notes",
    "Study Materials",
    "Assignments",
    "Other Academic Information"
]

EXAM_OPTIONS = [
    "All",
    "Internal Exams",
    "Model Exams",
    "Semester Exams",
    "Exam Timetable"
]

TIMETABLE_OPTIONS = [
    "All",
    "Class Timetable",
    "Exam Timetable",
    "Academic Schedule"
]

NOTICE_OPTIONS = [
    "All",
    "College Notices",
    "Department Notices",
    "Important Announcements"
]

EVENT_OPTIONS = [
    "All",
    "Department Events",
    "Other Department Events",
    "College Events"
]

COMPETITION_OPTIONS = [
    "All",
    "Department Competitions",
    "Inter-Department Competitions",
    "College Competitions",
    "Technical & Cultural Competitions"
]

RESULT_OPTIONS = [
    "All",
    "Prize Winners",
    "Winning Department",
    "Individual Achievements",
    "Awards & Recognitions"
]

LEARNING_OPTIONS = [
    "All",
    "Seminars",
    "Workshops",
    "Training Programs",
    "Career & Skill Programs"
]

SKILL_OPTIONS = [
    "All",
    "Technical Skills",
    "Communication Skills",
    "Career Skills",
    "Other Skills / Courses"
]

MANAGE_OPTIONS = [
    "Academic Notes",
    "Exam Timetables",
    "Class Timetables",
    "Notices & Announcements",
    "Events & Activities",
    "Seminars & Workshops",
    "Competitions & Results",
    "Student Achievements"
]

UPDATE_OPTIONS = [
    "Academic Updates",
    "Exam Updates",
    "Events",
    "Workshops",
    "Seminars",
    "Competitions",
    "Important Notices",
    "All Updates"
]

PAYMENT_OPTIONS = [
    "Department Event Fees",
    "Cultural Event Fees",
    "Competition Fees",
    "Other College Fees"
]

PAYMENT_DETAILS_OPTIONS = [
    "Student Name / Register Number",
    "Amount Paid",
    "Payment Type",
    "Payment Status",
    "Date of Payment",
    "Person in Charge"
]

RESPONSIBILITY_OPTIONS = [
    "Staff",
    "Student",
    "Department Coordinator",
    "Other"
]

ADD_OPTIONS = [
    "Notes / Study Materials",
    "Timetables",
    "Notices",
    "Events",
    "Workshops / Seminars",
    "Competition Results",
    "Student Achievements",
    "Other Updates"
]


# ============================================================
# DESIGN
# ============================================================

st.markdown(
    f"""
    <style>

    .main-title {{
        color: {PURPLE};
        text-align: center;
        font-size: 46px;
        font-weight: bold;
        margin-top: 30px;
    }}

    .subtitle {{
        color: #777777;
        text-align: center;
        font-size: 19px;
        margin-bottom: 30px;
    }}

    .section-title {{
        color: {PURPLE};
        font-size: 30px;
        font-weight: bold;
        margin-bottom: 20px;
    }}

    .welcome-text {{
        color: #555555;
        text-align: center;
        font-size: 18px;
        line-height: 1.7;
    }}

    div.stButton > button {{
        border-radius: 10px;
        font-weight: bold;
        min-height: 45px;
    }}

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# SESSION STATE
# ============================================================

if "page" not in st.session_state:
    st.session_state.page = "welcome"

if "admin_logged_in" not in st.session_state:
    st.session_state.admin_logged_in = False

# Student details
if "student_name" not in st.session_state:
    st.session_state.student_name = ""

if "student_email" not in st.session_state:
    st.session_state.student_email = ""

if "student_register" not in st.session_state:
    st.session_state.student_register = ""

if "student_department" not in st.session_state:
    st.session_state.student_department = "Select Department"

if "student_year" not in st.session_state:
    st.session_state.student_year = "Select Year"

if "student_suggestions" not in st.session_state:
    st.session_state.student_suggestions = ""

# Staff details
if "staff_name" not in st.session_state:
    st.session_state.staff_name = ""

if "staff_id" not in st.session_state:
    st.session_state.staff_id = ""

if "staff_department" not in st.session_state:
    st.session_state.staff_department = ""

if "staff_designation" not in st.session_state:
    st.session_state.staff_designation = "Select Designation"

if "staff_suggestions" not in st.session_state:
    st.session_state.staff_suggestions = ""


# ============================================================
# FUNCTIONS
# ============================================================

def go_to(page):
    st.session_state.page = page


def get_checked(prefix, options):

    selected = []

    for option in options:

        key = f"{prefix}_{option}"

        if st.session_state.get(key, False):
            selected.append(option)

    if not selected:
        return "None"

    return ", ".join(selected)


# ============================================================
# SAVE RESPONSE TO EXCEL
# ============================================================

def save_response(response):

    # Create new Excel file
    if not DATA_FILE.exists():

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Responses"

        headers = list(response.keys())

        # Header row
        for column, header in enumerate(headers, start=1):

            sheet.cell(
                row=1,
                column=column,
                value=header
            )

        # First response
        for column, header in enumerate(headers, start=1):

            sheet.cell(
                row=2,
                column=column,
                value=response.get(header, "")
            )

        workbook.save(DATA_FILE)

    else:

        # Open existing Excel file
        workbook = load_workbook(DATA_FILE)

        sheet = workbook["Responses"]

        # Read existing headers
        headers = []

        for cell in sheet[1]:

            if cell.value is not None:
                headers.append(cell.value)

        # Add new columns if required
        for key in response.keys():

            if key not in headers:

                headers.append(key)

                sheet.cell(
                    row=1,
                    column=len(headers),
                    value=key
                )

        # Add response as new row
        new_row = sheet.max_row + 1

        for column, header in enumerate(headers, start=1):

            sheet.cell(
                row=new_row,
                column=column,
                value=response.get(header, "")
            )

        workbook.save(DATA_FILE)


# ============================================================
# WELCOME PAGE
# ============================================================

if st.session_state.page == "welcome":

    st.markdown(
        '<div class="main-title">💜 Campus Sphere</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="subtitle">'
        'Welcome to Campus Sphere'
        '</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        """
        <div class="welcome-text">

        Welcome to <b>Campus Sphere</b> — your college
        information and feedback portal.

        <br><br>

        Connect, discover and share information
        within your campus community.

        </div>
        """,
        unsafe_allow_html=True
    )

    st.write("")
    st.write("")
    st.write("")

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:

        if st.button(
            "💜 Enter the Sphere",
            type="primary",
            use_container_width=True
        ):

            go_to("user_type")
            st.rerun()


# ============================================================
# STUDENT / STAFF
# ============================================================

elif st.session_state.page == "user_type":

    st.markdown(
        '<div class="section-title">'
        'Choose Your Category'
        '</div>',
        unsafe_allow_html=True
    )

    st.write(
        "Please select whether you are a student or staff member."
    )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "👨‍🎓 Student",
            use_container_width=True
        ):

            go_to("student1")
            st.rerun()

    with col2:

        if st.button(
            "👩‍🏫 Staff",
            use_container_width=True
        ):

            go_to("staff1")
            st.rerun()

    st.write("")

    if st.button(
        "💜 Rewind",
        use_container_width=True
    ):

        go_to("welcome")
        st.rerun()


# ============================================================
# STUDENT PAGE 1
# ============================================================

elif st.session_state.page == "student1":

    st.markdown(
        '<div class="section-title">'
        '👨‍🎓 Student Information'
        '</div>',
        unsafe_allow_html=True
    )

    st.text_input(
        "Name *",
        key="student_name"
    )

    st.text_input(
        "Email ID",
        key="student_email"
    )

    st.text_input(
        "Register Number",
        key="student_register"
    )

    st.selectbox(
        "Department *",
        [
            "Select Department",
            "B.Sc Data Science",
            "BCA",
            "B.Com",
            "B.Com CA",
            "History",
            "Mathematics",
            "Other"
        ],
        key="student_department"
    )

    st.selectbox(
        "Year of Study *",
        [
            "Select Year",
            "1st Year",
            "2nd Year",
            "3rd Year"
        ],
        key="student_year"
    )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("user_type")
            st.rerun()

    with col2:

        if st.button(
            "💜 Discover",
            type="primary",
            use_container_width=True
        ):

            if not st.session_state.student_name.strip():

                st.error("Please enter your name.")

            elif (
                st.session_state.student_department
                == "Select Department"
            ):

                st.error("Please select your department.")

            elif (
                st.session_state.student_year
                == "Select Year"
            ):

                st.error("Please select your year of study.")

            else:

                go_to("student2")
                st.rerun()


# ============================================================
# STUDENT PAGE 2
# ============================================================

elif st.session_state.page == "student2":

    st.markdown(
        '<div class="section-title">'
        '📚 Academic Information'
        '</div>',
        unsafe_allow_html=True
    )

    st.subheader("Academic Information")

    cols = st.columns(2)

    for i, option in enumerate(ACADEMIC_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"academic_{option}"
            )

    st.subheader("Exam Information")

    cols = st.columns(2)

    for i, option in enumerate(EXAM_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"exam_{option}"
            )

    st.subheader("Timetable")

    cols = st.columns(2)

    for i, option in enumerate(TIMETABLE_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"timetable_{option}"
            )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("student1")
            st.rerun()

    with col2:

        if st.button(
            "💜 Discover",
            type="primary",
            use_container_width=True
        ):

            go_to("student3")
            st.rerun()


# ============================================================
# STUDENT PAGE 3
# ============================================================

elif st.session_state.page == "student3":

    st.markdown(
        '<div class="section-title">'
        '📢 College Activities & Updates'
        '</div>',
        unsafe_allow_html=True
    )

    st.subheader("Notices")

    cols = st.columns(2)

    for i, option in enumerate(NOTICE_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"notice_{option}"
            )

    st.subheader("Events")

    cols = st.columns(2)

    for i, option in enumerate(EVENT_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"event_{option}"
            )

    st.subheader("Competitions")

    cols = st.columns(2)

    for i, option in enumerate(COMPETITION_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"competition_{option}"
            )

    st.subheader("Results & Achievements")

    cols = st.columns(2)

    for i, option in enumerate(RESULT_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"result_{option}"
            )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("student2")
            st.rerun()

    with col2:

        if st.button(
            "💜 Discover",
            type="primary",
            use_container_width=True
        ):

            go_to("student4")
            st.rerun()


# ============================================================
# STUDENT PAGE 4
# ============================================================

elif st.session_state.page == "student4":

    st.markdown(
        '<div class="section-title">'
        '🎯 Learning, Skills & Suggestions'
        '</div>',
        unsafe_allow_html=True
    )

    st.subheader(
        "Seminars / Workshops / Learning"
    )

    cols = st.columns(2)

    for i, option in enumerate(LEARNING_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"learning_{option}"
            )

    st.subheader("Skills / Courses")

    cols = st.columns(2)

    for i, option in enumerate(SKILL_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"skill_{option}"
            )

    st.subheader("Suggestions")

    st.text_area(
        "Your suggestions / additional feedback",
        key="student_suggestions",
        height=150
    )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("student3")
            st.rerun()

    with col2:

        if st.button(
            "💜 Submit Response",
            type="primary",
            use_container_width=True
        ):

            response = {

                "Submitted At":
                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                "User Type":
                    "Student",

                "Name":
                    st.session_state.get(
                        "student_name",
                        ""
                    ),

                "Email ID":
                    st.session_state.get(
                        "student_email",
                        ""
                    ),

                "Register Number":
                    st.session_state.get(
                        "student_register",
                        ""
                    ),

                "Department":
                    st.session_state.get(
                        "student_department",
                        ""
                    ),

                "Year":
                    st.session_state.get(
                        "student_year",
                        ""
                    ),

                "Academic Information":
                    get_checked(
                        "academic",
                        ACADEMIC_OPTIONS
                    ),

                "Exam Information":
                    get_checked(
                        "exam",
                        EXAM_OPTIONS
                    ),

                "Timetable":
                    get_checked(
                        "timetable",
                        TIMETABLE_OPTIONS
                    ),

                "Notices":
                    get_checked(
                        "notice",
                        NOTICE_OPTIONS
                    ),

                "Events":
                    get_checked(
                        "event",
                        EVENT_OPTIONS
                    ),

                "Competitions":
                    get_checked(
                        "competition",
                        COMPETITION_OPTIONS
                    ),

                "Results":
                    get_checked(
                        "result",
                        RESULT_OPTIONS
                    ),

                "Seminars / Workshops / Learning":
                    get_checked(
                        "learning",
                        LEARNING_OPTIONS
                    ),

                "Skills / Courses":
                    get_checked(
                        "skill",
                        SKILL_OPTIONS
                    ),

                "Suggestions":
                    st.session_state.get(
                        "student_suggestions",
                        ""
                    )
            }

            save_response(response)

            st.success(
                "💜 Your response has been submitted successfully!"
            )


# ============================================================
# STAFF PAGE 1
# ============================================================

elif st.session_state.page == "staff1":

    st.markdown(
        '<div class="section-title">'
        '👩‍🏫 Staff Information'
        '</div>',
        unsafe_allow_html=True
    )

    st.text_input(
        "Staff Name *",
        key="staff_name"
    )

    st.text_input(
        "Staff ID *",
        key="staff_id"
    )

    st.text_input(
        "Department *",
        key="staff_department"
    )

    st.selectbox(
        "Designation *",
        [
            "Select Designation",
            "Assistant Professor",
            "Associate Professor",
            "Professor",
            "Head of Department",
            "Coordinator",
            "Other"
        ],
        key="staff_designation"
    )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("user_type")
            st.rerun()

    with col2:

        if st.button(
            "💜 Discover",
            type="primary",
            use_container_width=True
        ):

            if not st.session_state.staff_name.strip():

                st.error("Please enter staff name.")

            elif not st.session_state.staff_id.strip():

                st.error("Please enter Staff ID.")

            elif not st.session_state.staff_department.strip():

                st.error("Please enter department.")

            elif (
                st.session_state.staff_designation
                == "Select Designation"
            ):

                st.error("Please select designation.")

            else:

                go_to("staff2")
                st.rerun()


# ============================================================
# STAFF PAGE 2
# ============================================================

elif st.session_state.page == "staff2":

    st.markdown(
        '<div class="section-title">'
        '🗂️ Staff Information & Updates'
        '</div>',
        unsafe_allow_html=True
    )

    st.subheader("Information Managed")

    cols = st.columns(2)

    for i, option in enumerate(MANAGE_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"manage_{option}"
            )

    st.subheader("Regular Updates")

    cols = st.columns(2)

    for i, option in enumerate(UPDATE_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"update_{option}"
            )

    st.subheader("Payment Information")

    cols = st.columns(2)

    for i, option in enumerate(PAYMENT_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"payment_{option}"
            )

    st.subheader("Payment Details")

    cols = st.columns(2)

    for i, option in enumerate(PAYMENT_DETAILS_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"payment_details_{option}"
            )

    st.subheader("Responsibility")

    cols = st.columns(2)

    for i, option in enumerate(RESPONSIBILITY_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"responsibility_{option}"
            )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("staff1")
            st.rerun()

    with col2:

        if st.button(
            "💜 Discover",
            type="primary",
            use_container_width=True
        ):

            go_to("staff3")
            st.rerun()


# ============================================================
# STAFF PAGE 3
# ============================================================

elif st.session_state.page == "staff3":

    st.markdown(
        '<div class="section-title">'
        '📝 Additional Staff Information'
        '</div>',
        unsafe_allow_html=True
    )

    st.subheader(
        "Information Added / Updated"
    )

    cols = st.columns(2)

    for i, option in enumerate(ADD_OPTIONS):

        with cols[i % 2]:

            st.checkbox(
                option,
                key=f"add_{option}"
            )

    st.subheader(
        "Additional Features / Improvements"
    )

    st.text_area(
        "Suggestions / additional information",
        key="staff_suggestions",
        height=150
    )

    st.write("")

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "💜 Rewind",
            use_container_width=True
        ):

            go_to("staff2")
            st.rerun()

    with col2:

        if st.button(
            "💜 Submit Response",
            type="primary",
            use_container_width=True
        ):

            response = {

                "Submitted At":
                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                "User Type":
                    "Staff",

                "Name":
                    st.session_state.get(
                        "staff_name",
                        ""
                    ),

                "Staff ID":
                    st.session_state.get(
                        "staff_id",
                        ""
                    ),

                "Department":
                    st.session_state.get(
                        "staff_department",
                        ""
                    ),

                "Designation":
                    st.session_state.get(
                        "staff_designation",
                        ""
                    ),

                "Information Managed":
                    get_checked(
                        "manage",
                        MANAGE_OPTIONS
                    ),

                "Regular Updates":
                    get_checked(
                        "update",
                        UPDATE_OPTIONS
                    ),

                "Payment Information":
                    get_checked(
                        "payment",
                        PAYMENT_OPTIONS
                    ),

                "Payment Details":
                    get_checked(
                        "payment_details",
                        PAYMENT_DETAILS_OPTIONS
                    ),

                "Responsibility":
                    get_checked(
                        "responsibility",
                        RESPONSIBILITY_OPTIONS
                    ),

                "Information Added / Updated":
                    get_checked(
                        "add",
                        ADD_OPTIONS
                    ),

                "Suggestions":
                    st.session_state.get(
                        "staff_suggestions",
                        ""
                    )
            }

            save_response(response)

            st.success(
                "💜 Your response has been submitted successfully!"
            )


# ============================================================
# ADMIN PAGE
# ============================================================

elif st.session_state.page == "admin":

    st.markdown(
        '<div class="section-title">'
        '🔐 Private Response Viewer'
        '</div>',
        unsafe_allow_html=True
    )

    st.write(
        "Only the administrator can view submitted responses."
    )

    # --------------------------------------------------------
    # PASSWORD
    # --------------------------------------------------------

    if not st.session_state.admin_logged_in:

        password = st.text_input(
            "Admin Password",
            type="password"
        )

        if st.button(
            "🔓 Login",
            type="primary"
        ):

            if password == ADMIN_PASSWORD:

                st.session_state.admin_logged_in = True
                st.rerun()

            else:

                st.error(
                    "❌ Incorrect password."
                )

    # --------------------------------------------------------
    # ADMIN LOGGED IN
    # --------------------------------------------------------

    else:

        st.success(
            "🔓 Admin access granted."
        )

        if DATA_FILE.exists():

            try:

                df = pd.read_excel(
                    DATA_FILE,
                    engine="openpyxl"
                )

                st.write(
                    f"### Total Responses: {len(df)}"
                )

                st.dataframe(
                    df,
                    use_container_width=True,
                    height=600
                )

                # --------------------------------------------
                # ADMIN ONLY EXCEL DOWNLOAD
                # --------------------------------------------

                excel_buffer = BytesIO()

                with pd.ExcelWriter(
                    excel_buffer,
                    engine="openpyxl"
                ) as writer:

                    df.to_excel(
                        writer,
                        index=False,
                        sheet_name="Responses"
                    )

                excel_buffer.seek(0)

                st.download_button(
                    "📥 Download Excel Responses",
                    data=excel_buffer.getvalue(),
                    file_name="Campus_Sphere_Responses.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    )
                )

            except Exception:

                st.error(
                    "Unable to read the Excel response file."
                )

        else:

            st.info(
                "No responses have been submitted yet."
            )

        st.write("")

        if st.button("🔒 Logout"):

            st.session_state.admin_logged_in = False
            st.rerun()


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.markdown(
        "## 💜 Campus Sphere"
    )

    st.divider()

    if st.button(
        "🏠 Welcome",
        use_container_width=True
    ):

        go_to("welcome")
        st.rerun()

    st.divider()

    if st.button(
        "🔐 Admin / View Responses",
        use_container_width=True
    ):

        go_to("admin")
        st.rerun()
