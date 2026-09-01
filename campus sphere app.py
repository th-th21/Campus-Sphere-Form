import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os


# ============================================================
# CAMPUS SPHERE
# ============================================================

root = tk.Tk()
root.title("Campus Sphere")
root.geometry("1050x700")
root.minsize(950, 650)
root.configure(bg="#FAF7FF")


# ============================================================
# COLORS
# ============================================================

PURPLE = "#4B248F"
DARK_PURPLE = "#321567"
LIGHT_PURPLE = "#F4EEFF"
BORDER_PURPLE = "#D8C9EE"
WHITE = "#FFFFFF"
GREY = "#777777"
ROOT_BG = "#FAF7FF"


# ============================================================
# VARIABLES
# ============================================================

current_page = 0
user_type = ""

# Student details
name_var = tk.StringVar()
email_var = tk.StringVar()
register_var = tk.StringVar()
department_var = tk.StringVar()
year_var = tk.StringVar()

# Staff details
staff_name_var = tk.StringVar()
staff_id_var = tk.StringVar()
staff_department_var = tk.StringVar()
designation_var = tk.StringVar()

# Student choices
academic_vars = {}
exam_vars = {}
timetable_vars = {}
notice_vars = {}
event_vars = {}
competition_vars = {}
result_vars = {}
learning_vars = {}
skill_vars = {}

# Student suggestion
suggestion_text = None

# Staff choices
staff_manage_vars = {}
staff_update_vars = {}
staff_payment_vars = {}
staff_payment_details_vars = {}
staff_responsibility_vars = {}
staff_add_update_vars = {}

staff_suggestion_text = None


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clear_window():
    for widget in root.winfo_children():
        widget.destroy()


def create_header(parent, page_text):

    header = tk.Frame(parent, bg=WHITE)
    header.pack(fill="x", padx=25, pady=(12, 0))

    # Page label
    page_label = tk.Label(
        header,
        text=page_text,
        font=("Arial", 9, "bold"),
        bg=PURPLE,
        fg=WHITE,
        padx=10,
        pady=4
    )
    page_label.pack(anchor="nw")

    # Title
    title_frame = tk.Frame(header, bg=WHITE)
    title_frame.pack(fill="x")

    title = tk.Label(
        title_frame,
        text="🎓  CAMPUS SPHERE",
        font=("Arial", 24, "bold"),
        bg=WHITE,
        fg=PURPLE
    )
    title.pack(pady=(2, 0))

    subtitle = tk.Label(
        title_frame,
        text="Student Information & Preferences Form",
        font=("Arial", 10),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    subtitle.pack(pady=(0, 6))

    # Heart
    heart = tk.Label(
        title_frame,
        text="♡",
        font=("Arial", 25, "bold"),
        bg=WHITE,
        fg="#8B5BC7"
    )
    heart.place(relx=0.98, rely=0.45, anchor="center")

    # Line
    line = tk.Frame(
        header,
        height=2,
        bg=PURPLE
    )
    line.pack(fill="x", pady=(5, 10))


def create_card(parent):
    card = tk.Frame(
        parent,
        bg=WHITE,
        highlightbackground=BORDER_PURPLE,
        highlightthickness=1
    )
    card.pack(fill="x", padx=25, pady=4)
    return card


def create_question(parent, number, question):

    frame = tk.Frame(parent, bg=WHITE)
    frame.pack(fill="x", padx=12, pady=(7, 2))

    label = tk.Label(
        frame,
        text=f"{number}.  {question}",
        font=("Arial", 9, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE,
        anchor="w",
        justify="left",
        wraplength=900
    )
    label.pack(anchor="w")

    return frame


def create_checkbox_group(parent, variables, options, columns=4):

    frame = tk.Frame(parent, bg=WHITE)
    frame.pack(fill="x", padx=12, pady=3)

    for i, option in enumerate(options):

        var = tk.BooleanVar(value=False)
        variables[option] = var

        cb = tk.Checkbutton(
            frame,
            text=option,
            variable=var,
            font=("Arial", 9),
            bg=WHITE,
            activebackground=WHITE,
            fg="#222222",
            selectcolor=WHITE,
            anchor="w"
        )

        row = i // columns
        col = i % columns

        cb.grid(
            row=row,
            column=col,
            sticky="w",
            padx=8,
            pady=4
        )

    return frame


def all_checkbox_control(variables):

    if "All" not in variables:
        return

    all_var = variables["All"]

    def select_all(*args):

        value = all_var.get()

        for option, var in variables.items():

            if option != "All":
                var.set(value)

    all_var.trace_add("write", select_all)


def get_selected(variables):

    return ", ".join(
        option
        for option, var in variables.items()
        if var.get()
    )


def create_input(parent, label_text, variable, placeholder=""):

    row = tk.Frame(parent, bg=WHITE)
    row.pack(fill="x", padx=15, pady=5)

    label = tk.Label(
        row,
        text=label_text,
        font=("Arial", 9, "bold"),
        bg=WHITE,
        fg="#222222",
        width=25,
        anchor="w"
    )
    label.pack(side="left")

    entry = tk.Entry(
        row,
        textvariable=variable,
        font=("Arial", 9),
        width=42,
        relief="solid",
        bd=1
    )
    entry.pack(side="left", ipady=4)

    return entry


def create_combobox(parent, label_text, variable, values):

    row = tk.Frame(parent, bg=WHITE)
    row.pack(fill="x", padx=15, pady=5)

    label = tk.Label(
        row,
        text=label_text,
        font=("Arial", 9, "bold"),
        bg=WHITE,
        fg="#222222",
        width=25,
        anchor="w"
    )
    label.pack(side="left")

    combo = ttk.Combobox(
        row,
        textvariable=variable,
        values=values,
        state="readonly",
        width=39
    )
    combo.pack(side="left")

    return combo


# ============================================================
# NAVIGATION
# ============================================================

def navigation(parent, back=True, next_page=True, submit=False):

    # This frame stays at the bottom of the page.
    nav = tk.Frame(
        parent,
        bg=WHITE,
        height=60
    )

    nav.pack(
        side="bottom",
        fill="x",
        padx=25,
        pady=(5, 12)
    )

    nav.pack_propagate(False)

    # Back / Rewind
    if back:

        rewind_button = tk.Button(
            nav,
            text="⏪  Rewind 💜",
            font=("Arial", 9, "bold"),
            bg=WHITE,
            fg=PURPLE,
            activebackground=LIGHT_PURPLE,
            activeforeground=PURPLE,
            relief="solid",
            bd=1,
            padx=15,
            pady=7,
            cursor="hand2",
            command=lambda: show_page(current_page - 1)
        )

        rewind_button.pack(
            side="left",
            pady=8
        )

    # Progress
    progress = tk.Label(
        nav,
        text=get_progress_text(),
        font=("Arial", 9),
        bg=WHITE,
        fg=DARK_PURPLE
    )

    progress.pack(
        side="left",
        expand=True
    )

    # Submit
    if submit:

        submit_button = tk.Button(
            nav,
            text="✓  Submit 💜",
            font=("Arial", 9, "bold"),
            bg=PURPLE,
            fg=WHITE,
            activebackground=DARK_PURPLE,
            activeforeground=WHITE,
            relief="flat",
            padx=18,
            pady=7,
            cursor="hand2",
            command=submit_form
        )

        submit_button.pack(
            side="right",
            pady=8
        )

    # Discover
    elif next_page:

        discover_button = tk.Button(
            nav,
            text="Discover 💜  ➜",
            font=("Arial", 9, "bold"),
            bg=PURPLE,
            fg=WHITE,
            activebackground=DARK_PURPLE,
            activeforeground=WHITE,
            relief="flat",
            padx=18,
            pady=7,
            cursor="hand2",
            command=lambda: show_page(current_page + 1)
        )

        discover_button.pack(
            side="right",
            pady=8
        )


def get_progress_text():

    if user_type == "Student":

        if current_page == 1:
            return "20% Completed"
        elif current_page == 2:
            return "50% Completed"
        elif current_page == 3:
            return "75% Completed"
        elif current_page == 4:
            return "100% Completed"

    elif user_type == "Staff":

        if current_page == 5:
            return "33% Completed"
        elif current_page == 6:
            return "66% Completed"
        elif current_page == 7:
            return "100% Completed"

    return ""


# ============================================================
# PAGE CONTROL
# ============================================================

def show_page(page):

    global current_page

    current_page = page

    clear_window()

    if page == 0:
        welcome_page()

    # ---------------- STUDENT ----------------

    elif page == 1:
        student_page_one()

    elif page == 2:
        student_page_two()

    elif page == 3:
        student_page_three()

    elif page == 4:
        student_page_four()

    # ---------------- STAFF ----------------

    elif page == 5:
        staff_page_one()

    elif page == 6:
        staff_page_two()

    elif page == 7:
        staff_page_three()


# ============================================================
# WELCOME PAGE
# ============================================================

def welcome_page():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=15
    )

    # Header without page number
    header = tk.Frame(
        main,
        bg=WHITE
    )
    header.pack(
        fill="x",
        padx=25,
        pady=(30, 0)
    )

    title = tk.Label(
        header,
        text="🎓  CAMPUS SPHERE",
        font=("Arial", 28, "bold"),
        bg=WHITE,
        fg=PURPLE
    )
    title.pack()

    subtitle = tk.Label(
        header,
        text="Student Information & Preferences Form",
        font=("Arial", 12),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    subtitle.pack(pady=(3, 12))

    heart = tk.Label(
        header,
        text="♡",
        font=("Arial", 30, "bold"),
        bg=WHITE,
        fg="#8B5BC7"
    )
    heart.place(
        relx=0.98,
        rely=0.45,
        anchor="center"
    )

    line = tk.Frame(
        header,
        height=2,
        bg=PURPLE
    )
    line.pack(
        fill="x",
        pady=5
    )

    # Welcome content
    content = tk.Frame(
        main,
        bg=WHITE
    )
    content.pack(
        fill="both",
        expand=True
    )

    welcome = tk.Label(
        content,
        text="Welcome to Campus Sphere",
        font=("Arial", 24, "bold"),
        bg=WHITE,
        fg=PURPLE
    )
    welcome.pack(
        pady=(75, 10)
    )

    description = tk.Label(
        content,
        text="Your campus information, all in one place.",
        font=("Arial", 12),
        bg=WHITE,
        fg=GREY
    )
    description.pack()

    enter_button = tk.Button(
        content,
        text="Enter the Sphere",
        font=("Arial", 12, "bold"),
        bg=PURPLE,
        fg=WHITE,
        activebackground=DARK_PURPLE,
        activeforeground=WHITE,
        relief="flat",
        padx=35,
        pady=12,
        cursor="hand2",
        command=choose_user_type
    )
    enter_button.pack(
        pady=55
    )

    bottom_heart = tk.Label(
        main,
        text="♥",
        font=("Arial", 22),
        bg=WHITE,
        fg="#8B5BC7"
    )
    bottom_heart.place(
        relx=0.90,
        rely=0.90
    )


# ============================================================
# USER TYPE PAGE
# ============================================================

def choose_user_type():

    global current_page

    current_page = 0

    clear_window()

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=15
    )

    create_header(
        main,
        "Choose User Type"
    )

    title = tk.Label(
        main,
        text="Who are you?",
        font=("Arial", 20, "bold"),
        bg=WHITE,
        fg=PURPLE
    )
    title.pack(
        pady=(50, 10)
    )

    description = tk.Label(
        main,
        text="Select your user type to continue.",
        font=("Arial", 11),
        bg=WHITE,
        fg=GREY
    )
    description.pack()

    buttons = tk.Frame(
        main,
        bg=WHITE
    )
    buttons.pack(
        pady=50
    )

    student_button = tk.Button(
        buttons,
        text="🎓  Student",
        font=("Arial", 11, "bold"),
        bg=PURPLE,
        fg=WHITE,
        activebackground=DARK_PURPLE,
        activeforeground=WHITE,
        relief="flat",
        padx=35,
        pady=12,
        cursor="hand2",
        command=start_student
    )
    student_button.grid(
        row=0,
        column=0,
        padx=20
    )

    staff_button = tk.Button(
        buttons,
        text="👩‍🏫  Staff",
        font=("Arial", 11, "bold"),
        bg=WHITE,
        fg=PURPLE,
        activebackground=LIGHT_PURPLE,
        activeforeground=PURPLE,
        relief="solid",
        bd=1,
        padx=40,
        pady=12,
        cursor="hand2",
        command=start_staff
    )
    staff_button.grid(
        row=0,
        column=1,
        padx=20
    )


def start_student():

    global user_type
    user_type = "Student"

    show_page(1)


def start_staff():

    global user_type
    user_type = "Staff"

    show_page(5)


# ============================================================
# STUDENT PAGE 1
# DETAILS + USER TYPE
# ============================================================

def student_page_one():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Page 1 of 4"
    )

    section = tk.Label(
        main,
        text="🎓  Student Details",
        font=("Arial", 13, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    description = tk.Label(
        main,
        text="Please provide your basic information.",
        font=("Arial", 9),
        bg=WHITE,
        fg=GREY
    )
    description.pack(
        anchor="w",
        padx=48,
        pady=(0, 5)
    )

    card = create_card(main)

    create_input(
        card,
        "1.  Name (Required)",
        name_var,
        "Enter your name"
    )

    create_input(
        card,
        "2.  Email ID (Optional)",
        email_var,
        "Enter your email ID"
    )

    create_input(
        card,
        "3.  Register Number (Optional)",
        register_var,
        "Enter your register number"
    )

    create_combobox(
        card,
        "4.  Department (Required)",
        department_var,
        [
            "B.Sc Data Science",
            "BCA",
            "B.Com",
            "B.Com CA",
            "History",
            "Mathematics",
            "Other"
        ]
    )

    create_combobox(
        card,
        "5.  Year of Study (Required)",
        year_var,
        [
            "1st Year",
            "2nd Year",
            "3rd Year"
        ]
    )

    # User type display
    user_label = tk.Label(
        main,
        text="User Type: Student",
        font=("Arial", 9, "bold"),
        bg=LIGHT_PURPLE,
        fg=PURPLE,
        padx=12,
        pady=6
    )
    user_label.pack(
        anchor="w",
        padx=30,
        pady=5
    )

    navigation(
        main,
        back=False,
        next_page=True
    )


# ============================================================
# STUDENT PAGE 2
# QUESTIONS 6 - 10
# ============================================================

def student_page_two():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Page 2 of 4"
    )

    section = tk.Label(
        main,
        text="📚  Academic, Examination, Timetables, Notices & Events",
        font=("Arial", 12, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    # Question 6
    card1 = create_card(main)

    create_question(
        card1,
        6,
        "Which academic information would you like to access through Campus Sphere?"
    )

    create_checkbox_group(
        card1,
        academic_vars,
        [
            "All",
            "Class Notes",
            "Study Materials",
            "Assignments",
            "Other Academic Information"
        ],
        columns=5
    )

    all_checkbox_control(academic_vars)

    # Question 7
    card2 = create_card(main)

    create_question(
        card2,
        7,
        "Which examination-related information would you like to receive through Campus Sphere?"
    )

    create_checkbox_group(
        card2,
        exam_vars,
        [
            "All",
            "Internal Exams",
            "Model Exams",
            "Semester Exams",
            "Exam Timetable"
        ],
        columns=5
    )

    all_checkbox_control(exam_vars)

    # Question 8
    card3 = create_card(main)

    create_question(
        card3,
        8,
        "Which timetable and regular academic updates would you like to access?"
    )

    create_checkbox_group(
        card3,
        timetable_vars,
        [
            "All",
            "Class Timetable",
            "Exam Timetable",
            "Academic Schedule"
        ],
        columns=4
    )

    all_checkbox_control(timetable_vars)

    # Question 9
    card4 = create_card(main)

    create_question(
        card4,
        9,
        "Which notices and announcements would you like to receive?"
    )

    create_checkbox_group(
        card4,
        notice_vars,
        [
            "All",
            "College Notices",
            "Department Notices",
            "Important Announcements"
        ],
        columns=4
    )

    all_checkbox_control(notice_vars)

    # Question 10
    card5 = create_card(main)

    create_question(
        card5,
        10,
        "Which events and activities would you like to know about?"
    )

    create_checkbox_group(
        card5,
        event_vars,
        [
            "All",
            "Department Events",
            "Other Department Events",
            "College Events"
        ],
        columns=4
    )

    all_checkbox_control(event_vars)

    navigation(
        main,
        back=True,
        next_page=True
    )


# ============================================================
# STUDENT PAGE 3
# QUESTIONS 11 - 14
# ============================================================

def student_page_three():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Page 3 of 4"
    )

    section = tk.Label(
        main,
        text="🏆  Opportunities, Learning & Skills",
        font=("Arial", 12, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    # Question 11
    card1 = create_card(main)

    create_question(
        card1,
        11,
        "Which competitions would you like to receive information about?"
    )

    create_checkbox_group(
        card1,
        competition_vars,
        [
            "All",
            "Department Competitions",
            "Inter-Department Competitions",
            "College Competitions",
            "Technical & Cultural Competitions"
        ],
        columns=3
    )

    all_checkbox_control(competition_vars)

    # Question 12
    card2 = create_card(main)

    create_question(
        card2,
        12,
        "What event and competition results would you like to know?"
    )

    create_checkbox_group(
        card2,
        result_vars,
        [
            "All",
            "Prize Winners",
            "Winning Department",
            "Individual Achievements",
            "Awards & Recognitions"
        ],
        columns=3
    )

    all_checkbox_control(result_vars)

    # Question 13
    card3 = create_card(main)

    create_question(
        card3,
        13,
        "Which seminars, workshops and learning programs interest you?"
    )

    create_checkbox_group(
        card3,
        learning_vars,
        [
            "All",
            "Seminars",
            "Workshops",
            "Training Programs",
            "Career & Skill Programs"
        ],
        columns=3
    )

    all_checkbox_control(learning_vars)

    # Question 14
    card4 = create_card(main)

    create_question(
        card4,
        14,
        "Which skills or courses are you interested in learning or improving?"
    )

    create_checkbox_group(
        card4,
        skill_vars,
        [
            "All",
            "Technical Skills",
            "Communication Skills",
            "Career Skills",
            "Other Skills / Courses"
        ],
        columns=3
    )

    all_checkbox_control(skill_vars)

    navigation(
        main,
        back=True,
        next_page=True
    )


# ============================================================
# STUDENT PAGE 4
# QUESTION 15 + SUBMIT
# ============================================================

def student_page_four():

    global suggestion_text

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Page 4 of 4"
    )

    section = tk.Label(
        main,
        text="💌  Suggestions",
        font=("Arial", 13, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    question = tk.Label(
        main,
        text="15.  What other information, skills, courses or campus activities would you like to see in Campus Sphere?",
        font=("Arial", 10, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE,
        wraplength=850,
        justify="left"
    )
    question.pack(
        anchor="w",
        padx=40,
        pady=(12, 8)
    )

    suggestion_text = tk.Text(
        main,
        height=7,
        font=("Arial", 10),
        relief="solid",
        bd=1,
        wrap="word"
    )

    suggestion_text.pack(
        fill="x",
        padx=40,
        pady=5
    )

    thank_you = tk.Label(
        main,
        text="💌\nThank You!",
        font=("Arial", 18, "bold"),
        bg=WHITE,
        fg=PURPLE
    )
    thank_you.pack(
        pady=12
    )

    small_text = tk.Label(
        main,
        text="Your response helps us build a better Campus Sphere for you.",
        font=("Arial", 9),
        bg=WHITE,
        fg=GREY
    )
    small_text.pack()

    navigation(
        main,
        back=True,
        next_page=False,
        submit=True
    )


# ============================================================
# STAFF PAGE 1
# STAFF DETAILS
# ============================================================

def staff_page_one():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Staff Page 1 of 3"
    )

    section = tk.Label(
        main,
        text="👩‍🏫  Staff Details",
        font=("Arial", 13, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    description = tk.Label(
        main,
        text="Please provide your college-related staff information.",
        font=("Arial", 9),
        bg=WHITE,
        fg=GREY
    )
    description.pack(
        anchor="w",
        padx=48,
        pady=(0, 8)
    )

    card = create_card(main)

    create_input(
        card,
        "Staff Name (Required)",
        staff_name_var,
        "Enter staff name"
    )

    create_input(
        card,
        "Staff ID (Required)",
        staff_id_var,
        "Enter staff ID"
    )

    create_combobox(
        card,
        "Department (Required)",
        staff_department_var,
        [
            "B.Sc Data Science",
            "BCA",
            "B.Com",
            "B.Com CA",
            "History",
            "Mathematics",
            "Other"
        ]
    )

    create_combobox(
        card,
        "Designation (Required)",
        designation_var,
        [
            "Assistant Professor",
            "Associate Professor",
            "Professor",
            "Head of Department",
            "Coordinator",
            "Other"
        ]
    )

    user_label = tk.Label(
        main,
        text="User Type: Staff",
        font=("Arial", 9, "bold"),
        bg=LIGHT_PURPLE,
        fg=PURPLE,
        padx=12,
        pady=6
    )
    user_label.pack(
        anchor="w",
        padx=30,
        pady=8
    )

    navigation(
        main,
        back=False,
        next_page=True
    )


# ============================================================
# STAFF PAGE 2
# QUESTIONS 1 - 4
# ============================================================

def staff_page_two():

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Staff Page 2 of 3"
    )

    section = tk.Label(
        main,
        text="📋  Campus Management & Updates",
        font=("Arial", 12, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    # Question 1
    card1 = create_card(main)

    create_question(
        card1,
        1,
        "Which campus information would you like to update or manage?"
    )

    create_checkbox_group(
        card1,
        staff_manage_vars,
        [
            "Academic Notes",
            "Exam Timetables",
            "Class Timetables",
            "Notices & Announcements",
            "Events & Activities",
            "Seminars & Workshops",
            "Competitions & Results",
            "Student Achievements"
        ],
        columns=4
    )

    # Question 2
    card2 = create_card(main)

    create_question(
        card2,
        2,
        "Which updates should be regularly communicated to students?"
    )

    create_checkbox_group(
        card2,
        staff_update_vars,
        [
            "Academic Updates",
            "Exam Updates",
            "Events",
            "Workshops",
            "Seminars",
            "Competitions",
            "Important Notices",
            "All Updates"
        ],
        columns=4
    )

    # Question 3
    card3 = create_card(main)

    create_question(
        card3,
        3,
        "Which college-related payment information should be maintained?"
    )

    create_checkbox_group(
        card3,
        staff_payment_vars,
        [
            "Department Event Fees",
            "Cultural Event Fees",
            "Competition Fees",
            "Other College Fees"
        ],
        columns=4
    )

    # Question 4
    card4 = create_card(main)

    create_question(
        card4,
        4,
        "Which payment details should be maintained?"
    )

    create_checkbox_group(
        card4,
        staff_payment_details_vars,
        [
            "Student Name / Register Number",
            "Amount Paid",
            "Payment Type",
            "Payment Status",
            "Date of Payment",
            "Person in Charge"
        ],
        columns=3
    )

    navigation(
        main,
        back=True,
        next_page=True
    )


# ============================================================
# STAFF PAGE 3
# QUESTIONS 5 - 7
# ============================================================

def staff_page_three():

    global staff_suggestion_text

    main = tk.Frame(
        root,
        bg=WHITE
    )

    main.pack(
        fill="both",
        expand=True,
        padx=20,
        pady=12
    )

    create_header(
        main,
        "Staff Page 3 of 3"
    )

    section = tk.Label(
        main,
        text="📚  Responsibilities, Updates & Suggestions",
        font=("Arial", 12, "bold"),
        bg=WHITE,
        fg=DARK_PURPLE
    )
    section.pack(
        anchor="w",
        padx=25
    )

    # Question 5
    card1 = create_card(main)

    create_question(
        card1,
        5,
        "Who is responsible for the payment or activity?"
    )

    create_checkbox_group(
        card1,
        staff_responsibility_vars,
        [
            "Staff",
            "Student",
            "Department Coordinator",
            "Other"
        ],
        columns=4
    )

    # Question 6
    card2 = create_card(main)

    create_question(
        card2,
        6,
        "What information should staff be able to add or update?"
    )

    create_checkbox_group(
        card2,
        staff_add_update_vars,
        [
            "Notes / Study Materials",
            "Timetables",
            "Notices",
            "Events",
            "Workshops / Seminars",
            "Competition Results",
            "Student Achievements",
            "Other Updates"
        ],
        columns=4
    )

    # Question 7
    card3 = create_card(main)

    create_question(
        card3,
        7,
        "What additional features or improvements would you suggest for Campus Sphere?"
    )

    staff_suggestion_text = tk.Text(
        card3,
        height=5,
        font=("Arial", 9),
        relief="solid",
        bd=1,
        wrap="word"
    )

    staff_suggestion_text.pack(
        fill="x",
        padx=12,
        pady=7
    )

    navigation(
        main,
        back=True,
        next_page=False,
        submit=True
    )


# ============================================================
# SUBMIT FORM
# ============================================================

def submit_form():

    # --------------------------------------------------------
    # STUDENT VALIDATION
    # --------------------------------------------------------

    if user_type == "Student":

        if name_var.get().strip() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please enter your Name."
            )
            return

        if department_var.get() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please select your Department."
            )
            return

        if year_var.get() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please select your Year of Study."
            )
            return

        suggestion = ""

        if suggestion_text is not None:
            suggestion = suggestion_text.get(
                "1.0",
                tk.END
            ).strip()

        save_student_data(suggestion)

    # --------------------------------------------------------
    # STAFF VALIDATION
    # --------------------------------------------------------

    elif user_type == "Staff":

        if staff_name_var.get().strip() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please enter Staff Name."
            )
            return

        if staff_id_var.get().strip() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please enter Staff ID."
            )
            return

        if staff_department_var.get() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please select Department."
            )
            return

        if designation_var.get() == "":
            messagebox.showwarning(
                "Campus Sphere",
                "Please select Designation."
            )
            return

        suggestion = ""

        if staff_suggestion_text is not None:
            suggestion = staff_suggestion_text.get(
                "1.0",
                tk.END
            ).strip()

        save_staff_data(suggestion)


# ============================================================
# SAVE STUDENT DATA
# ============================================================

def save_student_data(suggestion):

    file_name = "Campus_Sphere_Responses.xlsx"

    headers = [
        "User Type",
        "Name",
        "Email ID",
        "Register Number",
        "Department",
        "Year of Study",
        "Staff ID",
        "Designation",

        "Student - Academic Information",
        "Student - Examination Information",
        "Student - Timetable & Academic Updates",
        "Student - Notices & Announcements",
        "Student - Events & Activities",
        "Student - Competitions",
        "Student - Event & Competition Results",
        "Student - Seminars, Workshops & Learning Programs",
        "Student - Skills / Courses",
        "Student - Suggestions",

        "Staff - Information Managed",
        "Staff - Regular Updates",
        "Staff - Payment Information",
        "Staff - Payment Details",
        "Staff - Responsibility",
        "Staff - Information Added / Updated",
        "Staff - Suggestions"
    ]

    data = [
        "Student",
        name_var.get(),
        email_var.get(),
        register_var.get(),
        department_var.get(),
        year_var.get(),
        "",
        "",

        get_selected(academic_vars),
        get_selected(exam_vars),
        get_selected(timetable_vars),
        get_selected(notice_vars),
        get_selected(event_vars),
        get_selected(competition_vars),
        get_selected(result_vars),
        get_selected(learning_vars),
        get_selected(skill_vars),
        suggestion,

        "",
        "",
        "",
        "",
        "",
        "",
        ""
    ]

    save_to_excel(
        file_name,
        headers,
        data
    )


# ============================================================
# SAVE STAFF DATA
# ============================================================

def save_staff_data(suggestion):

    file_name = "Campus_Sphere_Responses.xlsx"

    headers = [
        "User Type",
        "Name",
        "Email ID",
        "Register Number",
        "Department",
        "Year of Study",
        "Staff ID",
        "Designation",

        "Student - Academic Information",
        "Student - Examination Information",
        "Student - Timetable & Academic Updates",
        "Student - Notices & Announcements",
        "Student - Events & Activities",
        "Student - Competitions",
        "Student - Event & Competition Results",
        "Student - Seminars, Workshops & Learning Programs",
        "Student - Skills / Courses",
        "Student - Suggestions",

        "Staff - Information Managed",
        "Staff - Regular Updates",
        "Staff - Payment Information",
        "Staff - Payment Details",
        "Staff - Responsibility",
        "Staff - Information Added / Updated",
        "Staff - Suggestions"
    ]

    data = [
        "Staff",
        staff_name_var.get(),
        "",
        "",
        staff_department_var.get(),
        "",
        staff_id_var.get(),
        designation_var.get(),

        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",

        get_selected(staff_manage_vars),
        get_selected(staff_update_vars),
        get_selected(staff_payment_vars),
        get_selected(staff_payment_details_vars),
        get_selected(staff_responsibility_vars),
        get_selected(staff_add_update_vars),
        suggestion
    ]

    save_to_excel(
        file_name,
        headers,
        data
    )


# ============================================================
# EXCEL SAVE FUNCTION
# ============================================================

def save_to_excel(file_name, headers, data):

    try:

        if os.path.exists(file_name):

            workbook = load_workbook(file_name)
            sheet = workbook.active

        else:

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Campus Sphere Responses"

            sheet.append(headers)

        sheet.append(data)

        # Make columns wider
        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:

                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except:
                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                35
            )

        workbook.save(file_name)

        messagebox.showinfo(
            "Campus Sphere",
            "Your response has been saved successfully! 💜\n\n"
            "The information is saved in:\n"
            "Campus_Sphere_Responses.xlsx"
        )

        root.destroy()

    except PermissionError:

        messagebox.showerror(
            "Excel File Open",
            "Please close the Campus_Sphere_Responses.xlsx "
            "file and try submitting again."
        )

    except Exception as error:

        messagebox.showerror(
            "Error",
            f"Could not save the response.\n\n{error}"
        )


# ============================================================
# START APPLICATION
# ============================================================

show_page(0)

root.mainloop()